import gui_object as gui
import excel_manager as xl
import util
import sys
import pandas as pd
import openpyxl
import calendar
import os
import time
import re
import requests
from bs4 import BeautifulSoup as bs
from datetime import datetime
from datetime import timedelta
from PyQt5.QtWidgets import *
from PyQt5.QtWidgets import QApplication
from PyQt5.QtCore import *


class Frame(QWidget):
    _debugLogDic = dict() # '기타'로 분류될 것들을 모아두는 곳
    _menuActions = list()
    file_name = "무제"
    _logtext = ""


    def __init__(self):
        super().__init__()
        self.InitUI()
    def InitUI(self):
        # 1
        hbox1 = QHBoxLayout()
        self.label_start = QLabel("시작일", self)
        self.date_start = QDateEdit(self)
        self.date_start.setDate(QDate.currentDate())
        self.date_start.setCalendarPopup(True)
        self.date_start.dateChanged.connect(self.AdjustDate)
        self.date_start.setMinimumDate(datetime(2000,1,1))

        self.label_end = QLabel("종료일", self)
        self.date_end = QDateEdit(self)
        self.date_end.setDate(QDate.currentDate())
        self.date_end.setCalendarPopup(True)
        self.date_end.dateChanged.connect(self.AdjustDate)

        hbox1.addWidget(self.label_start)
        hbox1.addWidget(self.date_start)
        hbox1.addSpacing(10)
        hbox1.addWidget(self.label_end)
        hbox1.addWidget(self.date_end)
        hbox1.addStretch()

        # 2
        hbox2 = QHBoxLayout()
        self.total_button = QPushButton("올인원 추출")
        self.total_button.clicked.connect(self.TotalProcess)
        hbox2.addWidget(self.total_button)

        # 3
        hbox3 = QHBoxLayout()
        self.analy_button = QPushButton("팀장님 분석표 추출")
        self.analy_button.clicked.connect(self.WriteLeaderInfo)
        hbox3.addWidget(self.analy_button)

        vbox1 = QVBoxLayout()
        vbox1.addLayout(hbox1)
        vbox1.addLayout(hbox2)
        vbox1.addLayout(hbox3)

        self.setLayout(vbox1)
    # 날짜 자동 보정
    def AdjustDate(self):
        self.date_end.setMinimumDate(self.date_start.date())
    # 달 선택
    def ChooseMonth(self):
        month = 0

        items = ("1월", "2월", "3월","4월", "5월", "6월","7월", "8월", "9월","10월", "11월", "12월")
        item, okPressed = QInputDialog.getItem(self, "CPA를 계산할 달을 선택하세요","(현재 연도 기준입니다.)", items, 0, False)
        if okPressed and item:
            print(item)
            item = item.replace("월", "")
            month = int(item)
        return month
    # 팀장 (아이디, 이름) 리스트
    def GetUserList(self):
        headers = {'User-Agent': 'python-requests/2.22.0', 'Accept-Encoding': 'gzip, deflate', 'Accept': '*/*', 'Connection': 'keep-alive'}
        cookies = {'ckGroup':'112', 'ckDepth':'1', 'ckAdmin':'sounghoo4699', 'ckAdminpw':'%2AB01432CAC77CFDE84E8E9AFDFF2AFCB6569660C9', 'ckName':'%EB%AC%B8%EC%84%B1%ED%9B%84+%EC%A3%BC%EC%9E%84', 'ckFlag':'Y', 'LogGrade':'A'}
            
        url = 'http://academy.myvilpt.gethompy.com/manager/reception/order_team_status.php'
        r = requests.get(url = url, headers = headers, cookies = cookies)
        r.encoding = 'utf-8'
        text = r.text

        soup = bs(text, 'html.parser')
        selector = soup.find("select",{'name':'id'})
        all_option = selector.find_all("option")
    
        username_regex = "^([A-Za-z0-9]){4,20}"
        name_regex     = "^([가-힣]){3}"
    
        user_list = list()

        for option in all_option[1:]:
            temp = option.text.split('/')
            try:
                username = re.match(username_regex, temp[0])[0]
                name     = re.match(name_regex, temp[1].replace(' ', ''))[0]

                user = {'아이디' : username, '이름' : name}
                user_list.append(user)
            except:
                print("%s 분리하는 도중 에러" % option.text)
                pass
        return user_list
    # 달력 날짜 범위에 따른 데이터 다운로드
    def DownloadData_by_cal(self):
        """
        파일 다운로드 하는 파트
        """
        start = self.date_start.date().toPyDate()
        end = self.date_end.date().toPyDate()
        address = "http://academy.myvilpt.gethompy.com/manager/reception/call_xls.php?left_step=1&r_no=&Page=1&grade=&day_p=r_regdate&r_part2=&id_type=&r_group=&search_word=&search_field=&sDate={s_date}%2000:00:00&eDate={e_date}%2023:59:59&s_Date=&e_Date=&r_status_1=&r_status_2=&r_status_3=&r_status_4=&r_status_5=&r_status_6=&r_status_7=".format(s_date = start, e_date = end)
        self.file_name = "./" + str(start) + "~" + str(end)
        print(self.file_name)

        try:
            util.download(address, self.file_name + ".html")
        except Exception as e:
            print("파일 다운로드 중 에러")
            print(e.args)
    # 선택한 달에 따른 데이터 다운로드
    # 파일 이름 반환
    def DownloadData_by_month(self, month, id, name, path):
        address = "http://academy.myvilpt.gethompy.com/manager/reception/order_team_xls.php?left_step=&total_day=&Page=1&" \
                + "id={p_id}&name={p_name}&year={p_year}&month={p_month}".format(p_id = id, p_name = name, p_year = datetime.now().year, p_month = month)
        self.file_name = name
        try:
            print(util.download(address, self.file_name + '.html', path, True))
            print(address)
        except Exception as e:
            print("파일 다운로드 중 에러")
            print(e.args)
        return self.file_name + ".html"
    # 파일 열기 액션
    def openFile(self):
        self.xl_file = xl.choose_excel_file(self)[0]
        print(self.xl_file)
    """
    플레이스 관련 함수들
    """
    def Processing_place_order(self):
        output_df = pd.read_excel(self.xl_file, '플레이스')
        classify_name_list = output_df.iloc[1,1:].dropna()[:11]
        classify_name_list = classify_name_list.replace("\s", "", regex = True)

        day_list = output_df.iloc[3:35,0].dropna()
        day_valid_list = output_df.iloc[4:35,1].dropna()

        # TODO
        # 인제 NaN 체크는 이런 데이터프레임을 만들어서 하도록
        # 요일 표시가 NaN인 경우만 골라서 데이터프레임을 만듬
        day_df = pd.concat([day_list, day_valid_list], axis = 1)
        day_df = day_df[~day_df.iloc[:,0].isnull() & day_df.iloc[:,1].isnull()]

        df = pd.concat([self.date_list, self.call_name_list, self.work_valid_list], axis = 1)
        df = df[df[4] != '(,)']
        df = df.replace("\s", "", regex = True)

        # 구분 이름에 따라 DF 분류
        # 구분 이름에 해당하지 않는 것은 "기타" DF로 분류
        extracted_df_dict = dict()
        for name in classify_name_list:
            extracted_df = df[df[3].str.contains("[플레이스]" + name, regex = False)]
            if(not extracted_df.empty):
                extracted_df_dict[name] = extracted_df
                df.loc[df[3].str.contains("[플레이스]" + name, regex = False), 4] = "True"
        extracted_df_dict["기타"] = df[df[3].str.contains("[플레이스]", regex = False) & ~df[4].isin(["True"])]

        util.DicDebug(extracted_df_dict)
        self._logtext += str(extracted_df_dict)

        # 메인 작업 시작
        WEEK_LIST = ['월', '화', '수', '목', '금', '토', '일']
        wb = openpyxl.load_workbook(self.xl_file)
        ws = wb['플레이스']

        start_dt = self.date_start.date()
        end_dt   = self.date_end.date()

        # 요일 표시가 NaN인 데이터프레임을 루프 돌림
        for day in day_df.values:
            if(not (start_dt.day() <= int(day[0]) and int(day[0]) <= end_dt.day())):
                continue
            cur_day = datetime(start_dt.year(), start_dt.month(), int(day[0]))
            row_index = int(day[0]) + 5 # day는 1부터 시작
            ws['B' + str(row_index)] = WEEK_LIST[cur_day.weekday()]

            # 루프를 돌면서 다른 구분을 만날때마다
            # 기록시작
            col_index = 3 # C
            max_col = output_df.shape[1]
            while(col_index < max_col):
                col_str = util.colnum_string(col_index)
                if ws[col_str + str(3)].value is not None:
                    name = ws[col_str + str(3)].value # 구분의 이름
                    try:
                        temp_df = extracted_df_dict[name]
                        # 버그! 
                        # 엑셀 숫자에 ' 이것을 붙여야 인식이 된다??
                        # 숫자인식에 큰 에러사항!
                        # 숫자를 문자로도 정수로도 인식하란 말인가?
                        # 지금은 숫자를 문자로 인식하고 있음 ex) 01, 02 같은 걸 처리하기 위해
                        cnt     = temp_df[temp_df[0].isin([int(day[0])])].shape[0]

                        ws[col_str + str(row_index)] = cnt
                    except Exception as e:
                        pass
                col_index += 1
                

        ws["A1"] = "{m}월 플레이스 오더".format(y = start_dt.year(), m = start_dt.month())
        ws["A2"] = "{y}.{m}".format(y = start_dt.year(), m = start_dt.month())


            # 기타로 분류될 지역 표시
        df = extracted_df_dict["기타"]
        if not df.empty:
            self._debugLogDic["기타_플레이스"] = df

        wb.save(self.xl_file)
        print("저장 완료")
    """
    직영팀 일일 오더 함수
    """
    def Processing_direct_team(self):
        output_df = pd.read_excel(self.xl_file, '직영팀')
        classify_name_list = output_df.iloc[1,1:].dropna()[:11]
        classify_name_list = classify_name_list.replace("\s", "", regex = True)

        day_list = output_df.iloc[:,0].dropna()
        day_valid_list = output_df.iloc[:,1]

        # TODO
        # 인제 NaN 체크는 이런 데이터프레임을 만들어서 하도록
        day_df = pd.concat([day_list, day_valid_list], axis = 1)
        day_df = day_df[~day_df.iloc[:,0].isnull() & day_df.iloc[:,1].isnull()].loc[4:34]
        
        df = pd.concat([self.date_list, self.call_name_list, self.work_valid_list], axis = 1)
        df = df[df[4] != '(,)']
        df = df.replace("\s", "", regex = True)

        # 구분 이름에 따라 DF 분류
        # 구분 이름에 해당하지 않는 것은 "기타" DF로 분류
        extracted_df_dict = dict()
        for name in classify_name_list:
            extracted_df = df[df[3].str.contains("본사/%s" % name, regex = False) | df[3].str.contains("본사_%s" % name, regex = False)]
            if(not extracted_df.empty):
                extracted_df_dict[name] = extracted_df
                df.loc[df[3].str.contains("본사/%s" % name, regex = False) | df[3].str.contains("본사_%s" % name, regex = False), 4] = "True"
        extracted_df_dict["기타"] = df[(df[3].str.contains("본사/", regex = False) | df[3].str.contains("본사_", regex = False)) & ~df[4].isin(["True"])]

        기타 = extracted_df_dict["기타"]
        if "메인블로그1" in extracted_df_dict.keys():
            extracted_df_dict["메인블로그1"] = extracted_df_dict["메인블로그1"].append(기타[기타[3].str.contains("본사/본사번호5482")])
        else:
            extracted_df_dict["메인블로그1"] = 기타[기타[3].str.contains("본사/본사번호5482")]
            
        if "메인블로그3" in extracted_df_dict.keys():
            extracted_df_dict["메인블로그3"] = extracted_df_dict["메인블로그3"].append(기타[기타[3].str.contains("본사/HYUN1531")])
        else:
            extracted_df_dict["메인블로그3"] = 기타[기타[3].str.contains("본사/HYUN1531")]

        extracted_df_dict["기타"] = extracted_df_dict["기타"][~extracted_df_dict["기타"][3].str.contains("본사/본사번호5482")]
        extracted_df_dict["기타"] = extracted_df_dict["기타"][~extracted_df_dict["기타"][3].str.contains("본사/HYUN1531")]

        util.DicDebug(extracted_df_dict)
        self._logtext += str(extracted_df_dict)

        # 메인 작업 시작
        WEEK_LIST = ['월', '화', '수', '목', '금', '토', '일']
        wb = openpyxl.load_workbook(self.xl_file)
        ws = wb['직영팀']
        
        start_dt = self.date_start.date()
        end_dt   = self.date_end.date()

        for day in day_df.values:
            if(not (start_dt.day() <= int(day[0]) and int(day[0]) <= end_dt.day())):
                continue
            cur_day = datetime(start_dt.year(), start_dt.month(), int(day[0]))
            row_index = int(day[0]) + 5
            ws['B' + str(row_index)] = WEEK_LIST[cur_day.weekday()]

            # 루프를 돌면서 다른 구분을 만날때마다
            # 기록시작
            col_index = 3 # C
            max_col = output_df.shape[1]
            while(col_index < max_col):
                col_str = util.colnum_string(col_index)
                if ws[col_str + str(3)].value is not None:
                    name = ws[col_str + str(3)].value
                    try:
                        temp_df = extracted_df_dict[name]
                        cnt     = temp_df[temp_df[0].isin([int(day[0])])].shape[0]

                        ws[col_str + str(row_index)] = cnt
                    except Exception as e:
                        pass
                col_index += 1

        ws["A1"] = "{y}.{m}월 직영팀 일일 오더 현황".format(y = start_dt.year(), m = start_dt.month())
        ws["A2"] = "{y}.{m}".format(y = start_dt.year(), m = start_dt.month())

        # 기타로 분류될 지역 표시
        df = extracted_df_dict["기타"]
        if not df.empty:
            self._debugLogDic["기타_직영팀"] = df

        wb.save(self.xl_file)
        print("저장 완료")
    """
    월보장 함수
    """
    def Processing_month_order(self):
        output_df  = pd.read_excel(self.xl_file, '월보장')
        classify_name_list = output_df.iloc[1,1:].dropna()[:10]
        classify_name_list = classify_name_list.replace("\s", "", regex = True)

        day_list = output_df.iloc[3:35,0].dropna()
        day_valid_list = output_df.iloc[3:,1].dropna()

        # TODO
        # 인제 NaN 체크는 이런 데이터프레임을 만들어서 하도록
        day_df = pd.concat([day_list, day_valid_list], axis = 1)
        day_df = day_df[~day_df.iloc[:,0].isnull() & day_df.iloc[:,1].isnull()]
        
        df = pd.concat([self.date_list, self.call_name_list, self.work_valid_list], axis = 1)
        df = df[df[4] != '(,)']
        df = df.replace("\s", "", regex = True)

        # 구분 이름에 따라 DF 분류
        # 구분 이름에 해당하지 않는 것은 "기타" DF로 분류
        extracted_df_dict = dict()
        for name in classify_name_list:
            extracted_df = df[df[3].str.contains("월보장_%s" % name, regex = False)]
            if(not extracted_df.empty):
                extracted_df_dict[name] = extracted_df
                df.loc[df[3].str.contains("월보장_%s" % name, regex = False), 4] = "True"
        extracted_df_dict["기타"] = df[df[3].str.contains("월보장_", regex = False) & ~df[4].isin(["True"])]

        util.DicDebug(extracted_df_dict)
        self._logtext += str(extracted_df_dict)

        #TODO 월보장 메인 처리 과정
        WEEK_LIST = ['월', '화', '수', '목', '금', '토', '일']
        wb = openpyxl.load_workbook(self.xl_file)
        ws = wb['월보장']
        
        start_dt = self.date_start.date()
        end_dt   = self.date_end.date()

        for day in day_df.values:
            if(not (start_dt.day() <= int(day[0]) and int(day[0]) <= end_dt.day())):
                continue
            cur_day = datetime(start_dt.year(), start_dt.month(), int(day[0]))
            row_index = int(day[0]) + 5
            ws['B' + str(row_index)] = WEEK_LIST[cur_day.weekday()]

            # 루프를 돌면서 다른 구분을 만날때마다
            # 기록시작
            col_index = 3 # C
            max_col = output_df.shape[1]
            while(col_index < max_col):
                col_str = util.colnum_string(col_index)
                if ws[col_str + str(3)].value is not None:
                    name = ws[col_str + str(3)].value          
                    try:
                        temp_df = extracted_df_dict[name]
                        cnt     = temp_df[temp_df[0].isin([int(day[0])])].shape[0]

                        ws[col_str + str(row_index)] = cnt
                    except Exception as e:
                        pass
                col_index += 1

        ws["A1"] = "{m}월 월보장 오더".format(m = start_dt.month())
        ws["A2"] = "{y}.{m}".format(y = start_dt.year(), m = start_dt.month())

        # 기타로 분류될 지역 표시
        df = extracted_df_dict["기타"]

        if not df.empty:
            self._debugLogDic["기타_월보장"] = df

        wb.save(self.xl_file)
        print("저장 완료")

        return
    """
    CPA 함수
    """
    def Processing_cpa(self):
        output_df  = pd.read_excel(self.xl_file, 'CPA')
        call_name_df = output_df.iloc[:,1].dropna()
        call_name_df = call_name_df.iloc[2:13] # 구분 이름들이 담김 dataframe
        
        df = pd.concat([self.date_list, self.call_name_list, self.work_valid_list], axis = 1)
        df = df[df[4] != '(,)']

        day_list = output_df.iloc[1,3:34].dropna()
        day_valid_list = output_df.iloc[2,3:34].dropna()

        day_df = pd.concat([day_list, day_valid_list], axis = 1, sort = False)
        day_df = day_df[~day_df.iloc[:,0].isnull() & day_df.iloc[:,1].isnull()]

        extracted_df_dict = dict()
        for call_name in call_name_df:
            extracted_df = df[df[3].str.contains('마케터/' + call_name)]
            if(not extracted_df.empty):
                extracted_df_dict[call_name] = extracted_df
                df.loc[df[3].str.contains('마케터/' + call_name), 4] = "True"
        extracted_df_dict["기타"] = df[df[3].str.contains("마케터/") & ~df[4].isin(["True"])]

        util.DicDebug(extracted_df_dict)
        self._logtext += str(extracted_df_dict)

        # 메인 작업 시작
        WEEK_LIST = ['월', '화', '수', '목', '금', '토', '일']
        start_dt = self.date_start.date()
        end_dt   = self.date_end.date()

        wb = openpyxl.load_workbook(self.xl_file)
        ws = wb['CPA']

        for day in day_df.values:
            if(not (start_dt.day() <= int(day[0]) and int(day[0]) <= end_dt.day())):
                continue
            col_index = util.colnum_string(util.string_colnum("C") + int(day[0]))
            
            # 요일 갱신
            cur_day = datetime(start_dt.year(), start_dt.month(), int(day[0]))
            ws[col_index + "4"] = WEEK_LIST[cur_day.weekday()]

            # 루프 시작
            row_index = 5
            max_row = output_df.shape[0] - 2
            while(row_index < max_row):
                name = ws['B' + str(row_index)].value
                try:
                    temp_df = extracted_df_dict[name]
                    cnt = temp_df[temp_df[0].isin([int(day[0])])].shape[0]

                    ws[col_index + str(row_index)] = cnt
                except Exception as e:
                    pass
                row_index += 1

        ws['B1'] = "{year}.{month}월 CPA 수익".format(year = start_dt.year(), month = start_dt.month())
        ws['B2'] = "{year}.{month}".format(year = start_dt.year(), month = start_dt.month())
         
        # 기타로 분류될 지역 표시
        df = extracted_df_dict["기타"]
        if not df.empty:
            self._debugLogDic["기타_cpa"] = df

        wb.save(self.xl_file)
        print("저장 완료")

    """
    지역별 오더 함수
    """
    # 원하는 형태로 엑셀 변환
    def Processing_order(self):
        try:
            output_df = pd.read_excel(self.xl_file, '지역별오더')
        except:
            print("파일을 열 수 없습니다.")
            return

        classify_area_list = output_df.iloc[8, 3:].dropna()
        
        day_list = output_df.iloc[:, 0].dropna()
        day_valid_list = output_df.iloc[:,1]

        day_df = pd.concat([day_list, day_valid_list], axis = 1)
        day_df = day_df[~day_df.iloc[:,0].isnull() & day_df.iloc[:,1].isnull()].loc[9:39]


        df = pd.concat([self.date_list, self.area_list, self.work_valid_list], axis = 1)
        df = df[df[4] != '(,)']

        extracted_df_dict = dict()
        for area in classify_area_list:
            regex_exp = "([\s]%s[시,도,군,구]$|^%s[시,도,군,구]$|[\s]%s[시,도,군,구][\s]|^%s[시,도,군,구][\s]|" % (area, area, area, area) \
                      + "[\s]%s$|^%s$|[\s]%s[\s]|^%s[\s])" % (area, area, area, area)

            extracted_df = df[df[8].str.contains(regex_exp, regex = True)]
            if(not extracted_df.empty):
                extracted_df_dict[area] = extracted_df
                df.loc[df[8].str.contains(regex_exp, regex = True), 4] = "True"
        extracted_df_dict["기타"] = df[~df[4].isin(["True"])]
        
        util.DicDebug(extracted_df_dict)
        self._logtext += str(extracted_df_dict)

        WEEK_LIST = ['월', '화', '수', '목', '금', '토', '일']
        wb = openpyxl.load_workbook(self.xl_file)
        ws = wb['지역별오더']

        start_dt = self.date_start.date()
        end_dt   = self.date_end.date()

        for day in day_df.values:
            if(not (start_dt.day() <= int(day[0]) and int(day[0]) <= end_dt.day())):
                continue
            cur_day = datetime(start_dt.year(), start_dt.month(), int(day[0]))
            row_index = int(day[0]) + 10
            ws['B' + str(row_index)] = WEEK_LIST[cur_day.weekday()]

            col_index = 4 # 시작
            max_col = output_df.shape[1] # 끝

            sum_start = util.colnum_string(col_index) + str(row_index)
            sum_end   = util.colnum_string(max_col) + str(row_index)
            ws['C' + str(row_index)] = '=SUM({start}:{end})'.format(start = sum_start, end = sum_end)

            # 루프를 돌면서 다른 구분을 만날때마다
            # 기록시작
            while(col_index <= max_col):
                col_str = util.colnum_string(col_index)
                if ws[col_str + str(10)].value is not None:
                    area = ws[col_str + str(10)].value
                    try:
                        temp_df = extracted_df_dict[area]
                        cnt = temp_df[temp_df[0].isin([int(day[0])])].shape[0]

                        ws[col_str + str(row_index)] = cnt
                    except Exception as e:
                        pass
                col_index += 1

        ws["A3"] = "기준월: {y}.{m}월".format(y = start_dt.year(), m = start_dt.month())
     

        # 기타로 분류될 지역 표시
        df = extracted_df_dict["기타"]
        if not df.empty:
            self._debugLogDic["기타_지역별오더"] = df

        # 엑셀 파일 저장
        wb.save(self.xl_file)
        print("저장 완료")
    """
    애드워즈 함수
    """
    def Processing_adwars(self):
        try:
            output_df = pd.read_excel(self.xl_file, '애드워즈')
        except:
            print("파일을 열 수 없습니다.")
            return

        day_list = output_df.iloc[:, 1].dropna()
        day_valid_list = output_df.iloc[:,2]

        day_df = pd.concat([day_list, day_valid_list], axis = 1)
        day_df = day_df[~day_df.iloc[:,0].isnull() & day_df.iloc[:,1].isnull()].loc[1:39]

        df = pd.concat([self.date_list, self.call_name_list, self.work_valid_list], axis = 1)
        df = df[df[4] != '(,)']

        extracted_df_dict = dict()

        extracted_df = df[df[3].str.contains("애드워즈/", regex = False)]
        if(not extracted_df.empty):
            extracted_df_dict["애드워즈"] = extracted_df
            df.loc[df[3].str.contains("애드워즈/", regex = False), 4] = "True"
        extracted_df_dict["기타"] = df[df[3].str.contains("애드워즈/", regex = False) & ~df[4].isin(["True"])]

        util.DicDebug(extracted_df_dict)

        WEEK_LIST = ['월', '화', '수', '목', '금', '토', '일']
        wb = openpyxl.load_workbook(self.xl_file)
        ws = wb['애드워즈']
        start_dt = self.date_start.date()
        end_dt   = self.date_end.date()

        for day in day_df.values:
            if(start_dt.day() <= int(day[0]) and int(day[0]) <= end_dt.day()):
                cur_day = datetime(start_dt.year(), start_dt.month(), int(day[0]))
                row_index = int(day[0]) + 3 
                ws['C' + str(row_index)] = WEEK_LIST[cur_day.weekday()]

                col_index = 'E'
                try:
                    # 또 인식 안되네
                    temp_df = extracted_df_dict["애드워즈"]
                    cnt = temp_df[temp_df[0].isin([int(day[0])])].shape[0]              

                    ws[col_index + str(row_index)] = cnt
                except Exception as e:
                    pass

        ws["A1"] = "{y}.{m}월 직영팀 일일 오더 현황".format(y = start_dt.year(), m = start_dt.month())
        ws["A2"] = "{y}.{m}".format(y = start_dt.year(), m = start_dt.month())

        # 기타로 분류될 지역 표시
        df = extracted_df_dict["기타"]
        if not df.empty:
            self._debugLogDic["기타_애드워즈"] = df

        # 엑셀 파일 저장
        wb.save(self.xl_file)
        print("저장 완료")
    def TotalProcess(self):
        # Input
        self.DownloadData_by_cal()
        self.input_df  = xl.read_excel(self.file_name + ".html")[0]
        
        # 날짜 리스트
        self.date_list = pd.Series(util.extract_date(date).day for date in self.input_df[0][1:])
        self.date_list.index += 1
        # 콜네임 리스트
        self.call_name_list = self.input_df[3][1:]
        # 일이 취소되었는지?
        self.work_valid_list = self.input_df[4][1:]
        self.area_list       = self.input_df[8][1:] # 주소
        
        # Output
        self.openFile()

        print("-------------------------")
        print("지역별오더 추출 시작")
        print("-------------------------\n")
        self.Processing_order()
        print("-------------------------")
        print("cpa 추출 시작")
        print("-------------------------\n")
        self.Processing_cpa()
        print("-------------------------")
        print("월보장 추출 시작")
        print("-------------------------\n")
        self.Processing_month_order()
        print("-------------------------")
        print("직영팀 추출 시작")
        print("-------------------------\n")
        self.Processing_direct_team()
        print("-------------------------")
        print("플레이스 추출 시작")
        print("-------------------------\n")
        self.Processing_place_order()
        print("-------------------------")
        print("애드워즈 추출 시작")
        print("-------------------------\n")
        self.Processing_adwars()
        print("-------------------------")
        print("로그 저장중")
        print("-------------------------\n")
        self.saveLog()
        print("로그 저장 완료!")

        os.system('start excel.exe "%s"' % (self.xl_file ))
    def WriteLeaderInfo(self):
        user_list = self.GetUserList()
        month = self.ChooseMonth()
        self.openFile()

        dir_path = util.CreateDir("팀장님폴더")
        output_wb = openpyxl.load_workbook(self.xl_file)

        for userInfo in user_list:
            # 파일 다운로드
            f_path = dir_path + '/' +  self.DownloadData_by_month(month, userInfo['아이디'], userInfo['이름'], dir_path)

            # pandas로 html을 읽어서 데이터를 옮겨줘야함
            # html을 읽어서 xlsx로 변환하는 방법을 생각해봤으나
            # 영 좋지 않은 결과로 xlsx 변환됨

            try:              
                output_ws = output_wb[userInfo['이름']]
                # B5 ~ Q35
                monthrange = calendar.monthrange(datetime.now().year, month)
                output_ws = xl.ClearWorkbook(output_ws, ['B', 'Q'], [5, 36])
                output_ws = xl.HtmlValueToWorkbook(f_path, output_ws, ['B', 'Q'], [5, monthrange[1] + 5])
            except Exception as e:
                print("파일을 열 수 없습니다.")
                print(self.file_name)
                print(e.args)
                continue

        output_wb.save(self.xl_file)
        os.system('start excel.exe "%s"' % (self.xl_file ))
    # 로그 저장
    def saveLog(self):
        try:
            temp_keys = list(self._debugLogDic.keys())
            with pd.ExcelWriter("기타.xlsx", mode = 'w') as writer:
                self._debugLogDic[temp_keys[0]].to_excel(writer, sheet_name = temp_keys[0])

            for key in temp_keys[1:]:
                with pd.ExcelWriter("기타.xlsx", mode = 'a') as writer:
                    self._debugLogDic[key].to_excel(writer, sheet_name = key)
        except:
            print("기타 없습니다")
            pass

        with open("trace_log.txt", 'w') as f:
            f.write(self._logtext)
    def Debug(self):
        return
class MainGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.Init()
    # 메뉴 초기화
    def InitMenu(self):
        exitAction = gui.MakeAction('닫기', self, 'Ctrl+Q', '프로그램 닫기')
        exitAction.triggered.connect(qApp.quit)

        totalAction = gui.MakeAction('올인원', self, 'Ctrl+O', '올인원 추출하기')
        totalAction.triggered.connect(self.body.TotalProcess)

        teamLeaderAnalyAction = gui.MakeAction('팀장님 분석표', self, 'Ctrl+P', '팀장님 분석표 추출하기')
        teamLeaderAnalyAction.triggered.connect(self.body.WriteLeaderInfo)


        menubar    = self.menuBar()
        exitTab    = gui.AddMenu(menubar, '닫기', exitAction)
        processTab = gui.AddMenu(menubar, '추출', totalAction)
        processTab.addAction(teamLeaderAnalyAction)

    # 모든 초기화 작업
    def Init(self):
        self.body = Frame()
        self.setCentralWidget(self.body)
        self.resize(300, 300)
        self.show()
        self.InitMenu()

def main():
    app = QApplication(sys.argv)
    main_gui = MainGUI()
    sys.exit(app.exec_())

main()