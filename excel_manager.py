import pandas as pd
import openpyxl as xl
import util
import codecs
from bs4 import BeautifulSoup as bs
from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtCore import QObject
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.styles.borders import Border

def choose_excel_file(parent:QObject):
    fname = QFileDialog.getOpenFileName(parent, 
                                        caption = "엑셀 파일을 선택하세요",
                                        filter = "Excels (*.xlsx);;All files(*.*)")
    return fname

def read_excel(fname, sheet_name = ""):
    try:
        df = pd.read_excel(fname, sheet_name)
    except:
        df = pd.read_html(fname)

    return df
# 엑셀 파일 만들기
def make_excel(file_name, sheet, desc = ""):
    try:
        pd.to_excel(file_name, sheet_name = sheet)
    except Exception as e:
        print("엑셀 파일을 저장하는 도중 오류가 발생했습니다.")
        print(e.args)
    return True
# 엑셀 파일 읽기
def load_dataframe(file_name, sheet_name):
    df = pd.read_excel(file_name, sheet_name)

    return df
def EmptyRow(sheet, min_row = 1, max_row = 1):
    for row in sheet.iter_rows(min_row = start, max_row = end):
        for cell in row:
            cell.value = ""
    return True
def HtmlValueToWorkbook(f_path, output_ws, col_range, row_range):
    result = None

    htmlFile = codecs.open(f_path, 'r', 'utf-8')
    document = bs(htmlFile.read(), features='lxml')
    tds = document.find_all('td', {'class':'style10'})

    # Html 루프를 돌면서 정보를 옮기기 위한 변수들
    tds_index = 0
    rep = {' ' : '', '원' : '', ',' : ''}

    for x in range(row_range[0], row_range[1]):
        row_index = str(x)
        for y in range(util.string_colnum(col_range[0]), util.string_colnum(col_range[1])):
            col_index = util.colnum_string(y)

            value = tds[tds_index].text

            # 숫자 문자열은 int로 변환
            if value.isdecimal():
                value = int(value)
            # '원' 이 포함된 돈관련 숫자를 int로 변환
            elif '원' in value:     
                value = int(util.replace_all(value, rep))

            output_ws[col_index + row_index] = value
            tds_index += 1
    result = output_ws
    
    return result

def ClearWorkbook(output_ws, col_range, row_range):
    for x in range(row_range[0], row_range[1]):
        row_index = str(x)
        for y in range(util.string_colnum(col_range[0]), util.string_colnum(col_range[1])):
            col_index = util.colnum_string(y)
            output_ws[col_index + row_index] = ""

    result = output_ws
    return result